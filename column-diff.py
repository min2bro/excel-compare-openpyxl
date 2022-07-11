import openpyxl
from openpyxl import load_workbook


def find_common_sheets(sheets_in_template, sheets_in_input):
    """Returns common sheets between template and input file"""

    sheets_missing_in_input_file = ",".join(
        list(set(sheets_in_template) - set(sheets_in_input))
    )
    c5 = active_sheet.cell(row=2, column=1)
    c5.value = sheets_missing_in_input_file

    sheets_missing_in_template_file = ",".join(
        list(set(sheets_in_input) - set(sheets_in_template))
    )
    c6 = active_sheet.cell(row=2, column=2)
    c6.value = sheets_missing_in_template_file

    common_sheets = list(set(sheets_in_template).intersection(sheets_in_input))

    return common_sheets


def find_missing_columns(sheet, template_wb, input_wb, input_file):
    """Returns missing column in input and template file"""

    header_row_template = [
        c.value for c in next(template_wb[sheet].iter_rows(min_row=1, max_row=1))
    ]
    header_row_input = [
        c.value for c in next(input_wb[sheet].iter_rows(min_row=1, max_row=1))
    ]

    columns_missing_in_input_file = list(
        set(header_row_template) - set(header_row_input)
    )
    columns_missing_in_template_file = list(
        set(header_row_input) - set(header_row_template)
    )

    missing_elem_index_input_file = [
        i
        for i, e in enumerate(header_row_template)
        if e in columns_missing_in_input_file
    ]

    # write to result
    if len(columns_missing_in_input_file) > 0:
        sheet = input_wb[sheet]
        for index in missing_elem_index_input_file:
            sheet.insert_cols(index)
            cell = sheet.cell(row=1, column=index)
            cell.value = header_row_template[index]
        input_wb.save(inputfile)

    return (
        columns_missing_in_input_file,
        columns_missing_in_template_file,
        missing_elem_index_input_file,
    )


def write_missing_cols(active_sheet, missing_col, file_type="input"):
    """Writes missing column to result.xlsx"""

    for col in range(0, len(missing_col) - 1):
        if file_type == "input":
            cell = active_sheet.cell(row=col + 2, column=1)
            cell.value = columns_missing_in_input_file[col]
        else:
            cell = active_sheet.cell(row=col + 2, column=2)
            cell.value = columns_missing_in_template_file[col]


if __name__ == "__main__":

    # Path: Input and Template file
    templatefile = "C:\\Users\\Downloads\\TEMPLATE_devops_input_master.xltx"
    inputfile = "C:\\Users\\Downloads\\prpc_devops_input_master.xlsx"
    resultfile = "C:\\Personal\\xltx-xlsx-column\\results.xlsx"

    # create result spreadsheet
    wb = openpyxl.Workbook()
    active_sheet = wb.create_sheet("sheet_comparison")
    c1 = active_sheet.cell(row=1, column=1)
    c1.value = "Missing_sheets_in_input_file"
    c2 = active_sheet.cell(row=1, column=2)
    c2.value = "Missing_sheets_in_template_file"

    # Load Input and Template spreadsheet
    template_wb = load_workbook(templatefile, data_only=True)
    input_wb = load_workbook(inputfile, data_only=True)
    sheets_in_template = template_wb.sheetnames
    sheets_in_input = input_wb.sheetnames

    # common sheets between input and template file
    common_sheets = find_common_sheets(sheets_in_template, sheets_in_input)

    # loop over each sheet and find missing columns
    for sheet in common_sheets:
        active_sheet = wb.create_sheet(f"{sheet}")
        c1 = active_sheet.cell(row=1, column=1)
        c1.value = "Missing_columns_in_input_file"
        c2 = active_sheet.cell(row=1, column=2)
        c2.value = "Missing_columns_in_template_file"

        (
            columns_missing_in_input_file,
            columns_missing_in_template_file,
            missing_elem_index_input_file,
        ) = find_missing_columns(sheet, template_wb, input_wb, inputfile)

        write_missing_cols(active_sheet, columns_missing_in_input_file, "input")
        write_missing_cols(active_sheet, columns_missing_in_template_file, "template")

    # Save File
    wb.save(resultfile)